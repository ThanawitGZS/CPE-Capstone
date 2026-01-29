import os
import re
from typing import List, Dict, Optional, Tuple
import pandas as pd
from datetime import datetime
import pytz
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# ============================================
# CONFIG — กำหนดโฟลเดอร์ output ของคุณเอง
# ============================================
OUT_DIR = "./output"
os.makedirs(OUT_DIR, exist_ok=True)

TH = r"\u0E00-\u0E7F"
UNIT_WORDS_BASE = ["ชุด","ตัว","ลูก","อัน","กล่อง","ดวง","จุด"]
UNIT_WORDS_EX = sorted(set(UNIT_WORDS_BASE) | set(["เส้น","ชิ้น","คู่","ใบ","แผ่น"]))
UNIT_NOISE_MAP = [
    r"ชื้น", r"ชนชื้น", r"ชีน", r"ชิ้นน", r"ชื้่น", r"ซีน", r"มซน",
    r"ชน", r"คน", r"ขึ้น", r"ขึน", r"ขึ่น"
]
money_pat_loose = r"\d{1,3}(?:,\d{3})*(?:[.:]\d{2})"
qty_unit_pat    = rf"(?:ลง\s*)?(\d+(?:\.\d+)?)\s*({'|'.join(UNIT_WORDS_EX)})\b"

ITEM_COLS_BASE = [
    "รหัสสินค้าบริษัทที่ผลิต",
    "รหัสสินค้าของบริษัทสั่งซื้อ",
    "รายการ",
    "จำนวน",
    "หน่วย",
    "ราคาต่อหน่วย",
    "ส่วนลด %",
    "ราคารวม",
]

def _qty_to_int_str_or_blank(s: str) -> str:
    if s is None:
        return ""
    txt = str(s).strip().replace(",", "")
    m = re.search(r"\d+(?:\.\d+)?", txt)
    if not m:
        return ""
    try:
        val = int(float(m.group(0)))
        return str(val)
    except:
        return ""

def _pre_clean_line(ln: str) -> str:
    ln = ln.replace("|", " ")
    ln = re.sub(r"[–—•·●]+", " ", ln)
    ln = re.sub(r"\s{2,}", " ", ln).strip()
    for w in UNIT_NOISE_MAP:
        ln = re.sub(fr"\b{w}\b", "ชิ้น", ln)
    ln = re.sub(r"\b(\d+)\s+[0O]\s*(ชิ้น)\b", r"\1 \2", ln)
    ln = re.sub(r"\b(\d+)\s+\1\s*(ชิ้น)\b", r"\1 \2", ln)
    ln = re.sub(r"\s+(VW|TW|MW)\s+", " ", ln)
    ln = re.sub(r"(\d)\s*[:.]\s*(\d{2})(?!\d)", r"\1.\2", ln)
    ln = re.sub(r"(\d\.\d{2})\s*[.:](?!\d)", r"\1", ln)
    ln = re.sub(r"(\d)\s+\.(\s*\d{2})", r"\1.\2", ln)
    ln = re.sub(r"(\d{1,3})\s*%\b", r"\1%", ln)
    ln = re.sub(r"\bS\s*(ตัว|เส้น|ชิ้น|คู่)\b", r"5 \1", ln)
    return ln

def _looks_like_measurement(tok: str) -> bool:
    t = tok.lower()
    if re.fullmatch(r"\d+(?:\.\d+)?(mm|cm|m|v|w|hz|ah|a|kw|kva|mah|ml|l)", t): return True
    if re.fullmatch(r"\d+(?:x|×|/)\d+(?:x|×|/\d+)?(?:mm|cm|m)?", t): return True
    if re.fullmatch(r"\d+(?:-\d+)+", t): return True
    if re.fullmatch(r"\d+/?\d*\"?", t): return True
    return False

def _normalize_token(tok: str) -> str:
    tok = tok.strip(" ,.;:()[]{}\"'")
    m = re.search(r"[A-Za-z][A-Za-z0-9\-\_\/\.]*$", tok)
    return m.group(0) if m else ""

def _is_valid_sku_token(tok: str) -> bool:
    if not tok:
        return False
    if re.search(f"[{TH}]", tok):
        return False
    if not re.search(r"[A-Za-z]", tok):
        return False
    if not re.search(r"\d", tok):
        return False
    if not re.fullmatch(r"[A-Za-z][A-Za-z0-9\-\_\/\.]*", tok):
        return False
    if _looks_like_measurement(tok):
        return False
    return True

def _has_product_code_header(text: str) -> bool:
    t = text.lower()
    if re.search(r"product\s*(code|key)", t): return True
    if re.search(r"รหัส\s*สินค้า", text): return True
    return False

def _header_allows_code(all_lines: List[str]) -> bool:
    head = " ".join(all_lines[:80])
    return _has_product_code_header(head)

def _is_sku_first_doc(all_lines: List[str]) -> bool:
    hits = 0
    for ln in all_lines:
        z = ln.strip()
        if re.match(r"^\d+\s", z):
            continue
        first = z.split()[0] if z.split() else ""
        if first and _is_valid_sku_token(_normalize_token(first)):
            if re.search(money_pat_loose, z) and (re.search(qty_unit_pat, z) or len(re.findall(money_pat_loose, z))>=2):
                hits += 1
    return hits >= 2

def _stitch_wrapped_lines(lines: List[str]) -> List[str]:
    stitched = []
    i = 0
    n = len(lines)
    while i < n:
        cur = _pre_clean_line(lines[i])
        nxt = _pre_clean_line(lines[i+1]) if i+1 < n else ""
        def is_itemish(s: str) -> bool:
            return bool(re.search(money_pat_loose, s)) and (re.search(qty_unit_pat, s) or len(re.findall(money_pat_loose, s)) >= 2)
        if re.match(r"^\s*\d+\s+", cur) and not is_itemish(cur) and nxt:
            join = (cur + " " + nxt).strip()
            if is_itemish(join):
                stitched.append(join)
                i += 2
                continue
        stitched.append(cur)
        i += 1
    return stitched

def _pick_sku(tokens: List[str], allowed_positions: List[int]) -> Tuple[str, List[str]]:
    toks = tokens[:]
    for pos in allowed_positions:
        idx = pos - 1
        if 0 <= idx < len(toks):
            t = _normalize_token(toks[idx])
            if _is_valid_sku_token(t):
                new = toks[:idx] + toks[idx+1:]
                return t, new
    return "", tokens

def _is_candidate_item_line(ln: str) -> bool:
    if any(kw in ln for kw in [
        "รวมเป็นเงิน","รวมทั้งสิ้น","จำนวนเงินรวม","จํานวนเงินรวม",
        "NET TOTAL","GRAND TOTAL","TOTAL","ภาษีมูลค่าเพิ่ม"
    ]):
        return False
    has_money = re.search(money_pat_loose, ln) is not None
    has_qty   = re.search(qty_unit_pat, ln) is not None
    if has_money and has_qty:
        return True
    return len(re.findall(money_pat_loose, ln)) >= 2

def _parse_line(ln: str, header_allows_code: bool, sku_first_doc: bool) -> Optional[Dict[str,str]]:
    mnum = re.match(r"^\s*(\d{1,3})\s+(.*)$", ln)
    if mnum:
        body = mnum.group(2).strip()
        toks = body.split()
        sku, rest_toks = _pick_sku(toks, allowed_positions=[1,2,3,4] if header_allows_code else [])
        rest = " ".join(rest_toks).strip()
    else:
        toks = ln.split()
        sku, rest_toks = _pick_sku(toks, allowed_positions=[1] if sku_first_doc else [])
        rest = " ".join(rest_toks).strip()
    moneys = re.findall(money_pat_loose, rest)
    unit_price = (moneys[-2] if len(moneys) >= 2 else "").replace(":", ".")
    line_total = (moneys[-1] if len(moneys) >= 1 else "").replace(":", ".")
    disc = ""
    md = re.search(r"(\d{1,3})\s*%", rest)
    if md: disc = md.group(1)
    qmatches = list(re.finditer(qty_unit_pat, rest))
    qty, unit, desc = "", "", rest
    if qmatches:
        q    = qmatches[-1]
        qty  = _qty_to_int_str_or_blank(q.group(1))
        unit = q.group(2)
        desc = rest[:q.start()].strip()
    return {
        "รหัสสินค้าบริษัทที่ผลิต": "",
        "รหัสสินค้าของบริษัทสั่งซื้อ": sku,
        "รายการ": desc,
        "จำนวน": qty,
        "หน่วย": unit,
        "ราคาต่อหน่วย": unit_price.replace(",", ""),
        "ส่วนลด %": disc,
        "ราคารวม": line_total.replace(",", ""),
    }

def extract_items_from_lines(per_page_lines: dict) -> pd.DataFrame:
    all_lines: List[str] = []
    for p in sorted(per_page_lines.keys()):
        all_lines.extend(per_page_lines[p])
    all_lines = [ln for ln in all_lines if len(ln.strip()) >= 3]
    all_lines = _stitch_wrapped_lines(all_lines)
    header_allows = _header_allows_code(all_lines)
    sku_first     = _is_sku_first_doc(all_lines)
    items: List[Dict[str, str]] = []
    for ln in all_lines:
        if _is_candidate_item_line(ln):
            rec = _parse_line(ln, header_allows, sku_first)
            if rec and (rec["จำนวน"] or rec["ราคารวม"]):
                items.append(rec)
    df = pd.DataFrame(items, columns=ITEM_COLS_BASE)
    df.insert(0, "ลำดับ", range(1, len(df)+1))
    return df

def today_th_ddmmyyyy_be() -> str:
    tz = pytz.timezone("Asia/Bangkok")
    now = datetime.now(tz)
    return f"{now.day:02d}/{now.month:02d}/{now.year + 543}"

def build_form_xlsx(df_items: pd.DataFrame, out_dir: str):
    company_name = ""
    import_date = today_th_ddmmyyyy_be()
    po_name = ""
    def _to_float(s):
        s = str(s).replace(",", "").strip()
        try:
            return float(s) if s not in ("", "nan", "None") else 0.0
        except:
            return 0.0
    grand_total = float(df_items["ราคารวม"].map(_to_float).sum())
    cols = df_items.columns.tolist()
    N = len(cols)
    def pad(v): return v + [""]*(N - len(v))
    rows = []
    rows.append(pad(["ชื่อบริษัท", company_name]))
    rows.append(pad(["วันที่นำเข้า", import_date]))
    rows.append(pad(["ชื่อใบสั่งซื้อ", po_name]))
    rows.append(cols)
    for _, r in df_items.iterrows():
        rows.append([r[c] for c in cols])
    total_row = [""]*N
    label_col = cols.index("รายการ")
    total_col = cols.index("ราคารวม")
    total_row[label_col] = "จำนวนเงินรวมทั้งสิ้น"
    total_row[total_col] = f"{grand_total:.2f}"
    rows.append(total_row)
    form_path = os.path.join(out_dir, "DataImport.xlsx")
    pd.DataFrame(rows).to_excel(form_path, index=False, header=False)
    wb = load_workbook(form_path)
    ws = wb.active
    for col in ws.columns:
        try:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
        except:
            pass
    ws.cell(row=ws.max_row, column=label_col+1).font = Font(bold=True)
    sum_cell = ws.cell(row=ws.max_row, column=total_col+1)
    sum_cell.font = Font(bold=True)
    sum_cell.alignment = Alignment(horizontal="right")
    sum_cell.number_format = "0.00"
    wb.save(form_path)
    print("✅ Saved FORM XLSX:", form_path)

def process_ocr_result(per_page_lines: dict):
    df_items = extract_items_from_lines(per_page_lines)
    # items_path = os.path.join(OUT_DIR, "items_for_import.xlsx")
    # df_items.to_excel(items_path, index=False)
    # print("✅ Saved ITEMS XLSX:", items_path)
    build_form_xlsx(df_items, OUT_DIR)

# ================== ตัวอย่างการเรียกใช้งาน ==================
if __name__ == "__main__":
    # สมมุติว่าคุณมี dict จาก OCR แล้ว
    # per_page_lines = {"1": ["1 สายไฟ 10 ชิ้น 100.00", "รวมทั้งสิ้น 100.00"], "2": [...]}
    from sample_ocr_result import per_page_lines  # <- แทนที่ด้วยผลจริง
    process_ocr_result(per_page_lines)
